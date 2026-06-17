"""Script para generar los fixtures de padrón para tests.

Ejecutar una sola vez desde backend/ con:
    python tests/fixtures/padron/generate_fixtures.py
"""
import csv
import io
import pathlib

HERE = pathlib.Path(__file__).parent


def create_valid_xlsx() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "apellidos", "email", "comision", "regional"])
    ws.append(["Juan", "Pérez", "juan@test.com", "A", "Norte"])
    ws.append(["María", "García", "maria@test.com", "B", "Sur"])
    ws.append(["Carlos", "López", "carlos@test.com", "A", ""])
    ws.append(["Ana", "Martínez", "ana@test.com", "", ""])
    ws.append(["Pedro", "Sánchez", "pedro@test.com", "B", "Norte"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_valid_csv() -> bytes:
    lines = [
        "nombre,apellidos,email,comision",
        "Juan,Pérez,juan@test.com,A",
        "María,García,maria@test.com,B",
        "Carlos,López,carlos@test.com,A",
        "Ana,Martínez,ana@test.com,",
        "Pedro,Sánchez,pedro@test.com,B",
    ]
    return "\n".join(lines).encode("utf-8")


def create_xlsx_sin_email() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "apellidos", "legajo"])  # sin columna email
    ws.append(["Juan", "Pérez", "12345"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_xlsx_grande() -> bytes:
    """Genera un xlsx con >5000 filas de datos."""
    import openpyxl

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(["nombre", "apellidos", "email"])  # header
    for i in range(5001):
        ws.append([f"Nombre{i}", f"Apellido{i}", f"user{i}@test.com"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    (HERE / "padron_valido.xlsx").write_bytes(create_valid_xlsx())
    (HERE / "padron_valido.csv").write_bytes(create_valid_csv())
    (HERE / "padron_sin_email.xlsx").write_bytes(create_xlsx_sin_email())
    (HERE / "padron_grande.xlsx").write_bytes(create_xlsx_grande())
    print("Fixtures generados en", HERE)
